"""
驗證腳本：模擬 GraphRAG 雙路徑遍歷，輸出 5 個事故情境的人資端與勞工端清單。
用於上傳前快速檢查 xlsx 內容是否串得起來。
"""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent / "資料" / "XLSX"


def load(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    return [dict(zip(header, r)) for r in rows[1:]]


accidents = {r["事故ID"]: r for r in load(ROOT / "節點" / "01_事故.xlsx")}
laws = {r["條文ID"]: r for r in load(ROOT / "節點" / "02_法規條文.xlsx")}
duties = {r["義務ID"]: r for r in load(ROOT / "節點" / "03_企業義務.xlsx")}
benefits = {r["權益ID"]: r for r in load(ROOT / "節點" / "04_勞工權益.xlsx")}
docs = {r["文件ID"]: r for r in load(ROOT / "節點" / "05_文件.xlsx")}
penalties = {r["罰則ID"]: r for r in load(ROOT / "節點" / "06_罰則.xlsx")}
units = {r["單位ID"]: r for r in load(ROOT / "節點" / "07_受理單位.xlsx")}

R1 = load(ROOT / "關係" / "R1_事故適用法規.xlsx")
R2 = load(ROOT / "關係" / "R2_法規衍生義務.xlsx")
R3 = load(ROOT / "關係" / "R3_法規衍生權益.xlsx")
R4 = load(ROOT / "關係" / "R4_義務需文件.xlsx")
R5 = load(ROOT / "關係" / "R5_權益需文件.xlsx")


def laws_for(aid: str) -> list[str]:
    return [r["條文ID"] for r in R1 if r["事故ID"] == aid]


def duties_for(law_ids: list[str]) -> list[str]:
    out = []
    for r in R2:
        if r["條文ID"] in law_ids and r["義務ID"] not in out:
            out.append(r["義務ID"])
    return out


def benefits_for(law_ids: list[str]) -> list[str]:
    out = []
    for r in R3:
        if r["條文ID"] in law_ids and r["權益ID"] not in out:
            out.append(r["權益ID"])
    return out


def docs_for_duty(d_ids: list[str]) -> list[str]:
    out = []
    for r in R4:
        if r["義務ID"] in d_ids and r["文件ID"] not in out:
            out.append(r["文件ID"])
    return out


def docs_for_benefit(b_ids: list[str]) -> list[str]:
    out = []
    for r in R5:
        if r["權益ID"] in b_ids and r["文件ID"] not in out:
            out.append(r["文件ID"])
    return out


def penalty_for_duty(d_id: str) -> list[dict]:
    duty = duties.get(d_id)
    if not duty:
        return []
    # 透過義務的法定依據（D001→L007 等）由罰則表反查
    out = []
    # 罰則表 法定依據條文ID → 找出該條文的義務
    for p in penalties.values():
        # 這裡簡化：罰則對應的條文若在 D001 的關聯法規中，就視為相關
        related_laws = [r["條文ID"] for r in R2 if r["義務ID"] == d_id]
        if p["法定依據條文ID"] in related_laws:
            out.append(p)
    return out


for aid, acc in accidents.items():
    print("=" * 70)
    print(f"事故 {aid}：{acc['名稱']}（{acc['上位類別']} / {acc['法律分類']}）")
    print("=" * 70)
    law_ids = laws_for(aid)
    print(f"\n適用法規（{len(law_ids)}）：")
    for lid in law_ids:
        l = laws[lid]
        print(f"  • {lid} {l['法規']} {l['條號']} — {l['條文摘要']}")

    print("\n──── 人資端（企業義務）────")
    d_ids = duties_for(law_ids)
    for did in d_ids:
        d = duties[did]
        deadline = f"{d['申報期限_小時']} 小時" if d['申報期限_小時'] else d['申報期限_文字']
        print(f"  • {did} {d['義務名稱']}（{deadline}）")
        plist = penalty_for_duty(did)
        for p in plist:
            print(f"      ⚠ 罰則 {p['罰則ID']}：{p['罰名']} {p['罰鍰下限_元']:,}–{p['罰鍰上限_元']:,} 元")
    print("  應備文件：")
    for doc_id in docs_for_duty(d_ids):
        print(f"    - {doc_id} {docs[doc_id]['文件名稱']}")

    print("\n──── 勞工端（可請領給付）────")
    b_ids = benefits_for(law_ids)
    for bid in b_ids:
        b = benefits[bid]
        u = units.get(b["受理單位ID"], {})
        print(f"  • {bid} {b['給付名稱']}（時效 {b['時效_年']} 年｜申請至 {u.get('名稱', '')}）")
        print(f"      條件：{b['申請條件']}")
    print("  應備文件：")
    for doc_id in docs_for_benefit(b_ids):
        print(f"    - {doc_id} {docs[doc_id]['文件名稱']}")
    print()
