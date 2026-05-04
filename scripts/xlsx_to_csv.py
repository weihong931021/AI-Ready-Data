"""
把 XLSX/節點 與 XLSX/關係 下所有檔案轉成 CSV（UTF-8）。
平台 Data Importer 用 CSV 而非 xlsx。
"""
import csv
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent / "資料"
SRC_BASE = ROOT / "XLSX"
DST_BASE = ROOT / "CSV"

for sub in ["節點", "關係"]:
    src = SRC_BASE / sub
    dst = DST_BASE / sub
    dst.mkdir(parents=True, exist_ok=True)
    for xlsx in sorted(src.glob("*.xlsx")):
        wb = load_workbook(xlsx, data_only=True)
        ws = wb.active
        out = dst / f"{xlsx.stem}.csv"
        with out.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        print(f"  ✓ {out.relative_to(ROOT)}")

print("\n所有 CSV 已產生（UTF-8 with BOM）。")
